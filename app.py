import pyautogui
import time
from openpyxl import load_workbook
import pywhatkit as pwk

wb = load_workbook("contacts.xlsx")
ws = wb.active

for i in range(2,4):
    contact_number = f"+{str(ws.cell(row=i, column=2).value)}"
    contact_name = ws.cell(row=i, column=1).value
    pwk.sendwhatmsg_instantly(str(contact_number),
						f'''Hey {contact_name}! How's it going?''')
    time.sleep(2)
    pyautogui.typewrite(["enter"])
    time.sleep(2)  